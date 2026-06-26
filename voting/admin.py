from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction
import pandas as pd
from io import BytesIO
from .models import Voter, Vote
from .forms import ExcelImportForm


@admin.register(Voter)
class VoterAdmin(admin.ModelAdmin):
    list_display = ['voter_no', 'full_name', 'house', 'pin_generated']
    list_filter = ['house']
    search_fields = ['voter_no', 'full_name', 'house']
    list_per_page = 50
    fields = ['voter_no', 'full_name', 'house', 'pin_status']
    readonly_fields = ['pin_status']

    actions = ['delete_all_voters', 'generate_and_export_selected_pins']

    change_list_template = "admin/voters/voter_changelist.html"

    @admin.display(boolean=True, description='PIN Generated')
    def pin_generated(self, obj):
        return bool(obj.pin)

    @admin.display(description='PIN Status')
    def pin_status(self, obj):
        if obj and obj.pin:
            return "Generated"
        return "Not Generated"

    @admin.action(description="Generate & Export PINs (Overwrite)")
    def generate_and_export_selected_pins(self, request, queryset):
        import random
        import string
        from io import BytesIO
        import pandas as pd
        from django.http import HttpResponse

        with transaction.atomic():
            voters = list(queryset)
            for voter in voters:
                voter.pin = "".join(random.choices(string.digits, k=6))
            Voter.objects.bulk_update(voters, ['pin'])

        # Optimize dataframe construction to prevent instantiating/looping model instances
        selected_ids = [v.id for v in voters]
        voters_data = Voter.objects.filter(id__in=selected_ids).values('voter_no', 'full_name', 'house', 'pin').order_by('full_name')
        df = pd.DataFrame(list(voters_data))
        df = df[['voter_no', 'full_name', 'house', 'pin']]

        from django.utils import timezone
        generation_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=2, sheet_name='Selected Voter PINs')

            workbook = writer.book
            worksheet = writer.sheets['Selected Voter PINs']
            from openpyxl.styles import Font, PatternFill

            # Write generation time at the top
            worksheet.cell(row=1, column=1, value=f"Generated on: {generation_time}")
            worksheet.cell(row=1, column=1).font = Font(italic=True, size=10)

            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[3]:
                cell.fill = header_fill
                cell.font = header_font

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 15

        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=selected_voter_pins.xlsx'
        return response

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='voters_voter_import_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template), name='voters_voter_download_template'),
            path('export-voter-pins/', self.admin_site.admin_view(self.export_voter_pins), name='voters_voter_export_pins'),
        ]
        return custom_urls + urls

    def export_voter_pins(self, request):
        """Generate PINs for voters missing them, and export all voters with PINs to Excel."""
        import random
        import string
        from io import BytesIO
        import pandas as pd
        from django.http import HttpResponse
        from django.db.models import Q

        # Generate PINs for voters missing them
        missing_voters = Voter.objects.filter(Q(pin__isnull=True) | Q(pin=""))
        if missing_voters.exists():
            with transaction.atomic():
                voters_list = list(missing_voters)
                for voter in voters_list:
                    voter.pin = "".join(random.choices(string.digits, k=6))
                Voter.objects.bulk_update(voters_list, ['pin'])

        # Export ALL voters using values() to prevent model instantiation and minimize memory usage
        voters_data = Voter.objects.values('voter_no', 'full_name', 'house', 'pin').order_by('full_name')
        df = pd.DataFrame(list(voters_data))
        df = df[['voter_no', 'full_name', 'house', 'pin']]

        from django.utils import timezone
        generation_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=2, sheet_name='Voter PINs')

            workbook = writer.book
            worksheet = writer.sheets['Voter PINs']
            from openpyxl.styles import Font, PatternFill

            # Write generation time at the top
            worksheet.cell(row=1, column=1, value=f"Generated on: {generation_time}")
            worksheet.cell(row=1, column=1).font = Font(italic=True, size=10)

            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[3]:
                cell.fill = header_fill
                cell.font = header_font

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 15

        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=voter_pins.xlsx'
        return response

    def import_excel(self, request):
        """Handle Excel/CSV file upload and import voters using pandas."""
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                overwrite = form.cleaned_data['overwrite_existing']

                # Read column mapping and header row from POST params
                col_voter_no = request.POST.get('col_voter_no', '').strip()
                col_full_name = request.POST.get('col_full_name', '').strip()
                col_house = request.POST.get('col_house', '').strip()
                header_row = int(request.POST.get('header_row', '0').strip() or '0')

                try:
                    file_name = excel_file.name.lower()

                    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls') or file_name.endswith('.csv')):
                        messages.error(request, "Please upload a valid Excel (.xlsx, .xls) or CSV file")
                        return redirect('.')

                    if excel_file.size > 10 * 1024 * 1024:
                        messages.error(request, "File size too large. Maximum size is 10MB.")
                        return redirect('.')

                    file_content = BytesIO(excel_file.read())

                    try:
                        if file_name.endswith('.csv'):
                            df = pd.read_csv(file_content, header=header_row)
                        else:
                            try:
                                df = pd.read_excel(file_content, header=header_row, engine='openpyxl')
                            except:
                                file_content.seek(0)
                                df = pd.read_excel(file_content, header=header_row, engine='xlrd')
                    except Exception as e:
                        messages.error(request, f"Could not read file: {str(e)}. Please ensure it's a valid Excel or CSV file.")
                        return redirect('.')

                    # Apply column mapping if provided
                    rename_map = {}
                    if col_voter_no and col_voter_no in df.columns:
                        rename_map[col_voter_no] = 'voter_no'
                    if col_full_name and col_full_name in df.columns:
                        rename_map[col_full_name] = 'full_name'
                    if col_house and col_house in df.columns:
                        rename_map[col_house] = 'house'
                    if rename_map:
                        df = df.rename(columns=rename_map)

                    # Validate required columns
                    required_columns = ['voter_no', 'full_name']
                    missing_columns = [col for col in required_columns if col not in df.columns]

                    if missing_columns:
                        messages.error(
                            request,
                            f"Missing required columns after mapping: {', '.join(missing_columns)}. "
                            f"Please map your file columns to voter_no and full_name."
                        )
                        return redirect('.')

                    # Clean data
                    df = df.fillna('')
                    df['voter_no'] = df['voter_no'].astype(str).str.strip()
                    df['full_name'] = df['full_name'].astype(str).str.strip()

                    if 'house' in df.columns:
                        df['house'] = df['house'].astype(str).str.strip()
                    else:
                        df['house'] = ''

                    # Remove empty rows
                    df = df[df['voter_no'] != '']
                    df = df[df['full_name'] != '']

                    if len(df) == 0:
                        messages.warning(request, "No valid data found in file")
                        return redirect('.')

                    created_count = 0
                    updated_count = 0
                    skipped_count = 0
                    errors = []

                    with transaction.atomic():
                        for index, row in df.iterrows():
                            try:
                                voter_no = row['voter_no']
                                name = row['full_name']
                                house = row.get('house', '')

                                if not voter_no or not name:
                                    errors.append(f"Row {index + 2}: Missing voter_no or full_name")
                                    skipped_count += 1
                                    continue

                                voter, created = Voter.objects.get_or_create(
                                    voter_no=voter_no,
                                    defaults={'full_name': name, 'house': house}
                                )

                                if created:
                                    created_count += 1
                                elif overwrite:
                                    voter.full_name = name
                                    voter.house = house
                                    voter.save(update_fields=['full_name', 'house'])
                                    updated_count += 1
                                else:
                                    skipped_count += 1

                            except Exception as e:
                                errors.append(f"Row {index + 2}: {str(e)}")
                                skipped_count += 1

                    if errors:
                        for error in errors[:10]:
                            messages.warning(request, error)
                        if len(errors) > 10:
                            messages.warning(request, f"... and {len(errors) - 10} more errors")

                    summary = f"Import complete: {created_count} created"
                    if updated_count:
                        summary += f", {updated_count} updated"
                    if skipped_count:
                        summary += f", {skipped_count} skipped"
                    messages.success(request, summary)
                    return redirect('..')

                except Exception as e:
                    messages.error(request, f"Error processing file: {str(e)}")
                    return redirect('.')
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'title': 'Import Voters from Excel/CSV',
            'site_title': 'Import Voters',
            'site_header': admin.site.site_header,
            'has_permission': True,
        }
        return render(request, 'admin/voters/import_excel.html', context)

    def download_template(self, request):
        """Download Excel template for voter import using pandas"""
        # Create sample data
        data = {
            'voter_no': ['V001', 'V002', 'V003'],
            'full_name': ['John Doe', 'Jane Smith', 'Mary Johnson'],
            'house': ['AGAKHAN', 'AFRICA', 'KAKUNGULU']
        }

        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Voters Template')

            # Access the workbook to style headers
            workbook = writer.book
            worksheet = writer.sheets['Voters Template']

            # Style headers
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Adjust column widths
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 20

        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=voters_template.xlsx'

        return response

    def delete_all_voters(self, request, queryset):
        """Delete all voters (use with caution)"""
        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"{count} voters deleted.", messages.WARNING)

    delete_all_voters.short_description = "Delete selected voters"


@admin.register(Vote)
class VoteAdmin(admin.ModelAdmin):
    list_display = ['voter', 'post', 'candidate', 'timestamp']
    list_filter = ['post', 'timestamp']
    search_fields = ['voter__full_name', 'voter__voter_no', 'candidate__name']
    readonly_fields = ['voter', 'post', 'candidate', 'timestamp']
    date_hierarchy = 'timestamp'

    def has_add_permission(self, request):
        # Prevent manual vote addition through admin
        return False

    def has_change_permission(self, request, obj=None):
        # Prevent vote modification through admin
        return False
