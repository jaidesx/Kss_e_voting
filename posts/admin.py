from django.contrib import admin
from django.shortcuts import redirect
from django.urls import path
from django.http import HttpResponse
from django.db.models import Count
from io import BytesIO
import openpyxl
from .models import Post, EligibleHouse, Election
from voting.models import Vote


class EligibleHouseInline(admin.TabularInline):
    model = EligibleHouse
    extra = 1
    verbose_name = 'Eligible House'
    verbose_name_plural = 'Eligible Houses (leave empty = open to all)'


def generate_results_excel(elections):
    """
    Generate an Excel workbook with results for the given elections.
    Returns a BytesIO object containing the xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.utils import timezone
    from voting.models import Voter, Vote

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    generation_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

    # Premium Color Palette & Typography
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_candidate = Font(name="Calibri", size=10)
    font_winner = Font(name="Calibri", size=10, bold=True, color="375623")
    font_stats_lbl = Font(name="Calibri", size=9, bold=True, color="595959")
    font_stats_val = Font(name="Calibri", size=10, bold=True, color="000000")

    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_subtitle = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    fill_section = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_winner = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
    fill_stats = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    def style_range(ws, start_row, start_col, end_row, end_col, font=None, fill=None, alignment=None, border=None):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border

    for election in elections:
        # Clean title for sheet name (max 31 chars, openpyxl rules)
        clean_title = "".join(c for c in election.title if c.isalnum() or c in " -_")[:30].strip()
        if not clean_title:
            clean_title = f"Election {election.id}"
        ws = wb.create_sheet(title=clean_title)

        # Explicitly make gridlines visible
        ws.views.sheetView[0].showGridLines = True

        # 1. Title Banner
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')

        # Style all cells in the title merged ranges
        style_range(ws, 1, 1, 1, 7, font=font_title, fill=fill_title, alignment=align_center)
        style_range(ws, 2, 1, 2, 7, font=font_subtitle, fill=fill_subtitle, alignment=align_center)

        ws['A1'].value = "KSS STUDENT BALLOT ELECTION RESULTS"
        ws['A2'].value = f"Election: {election.title} {'(Demo)' if election.is_demo else ''}"

        # 2. Statistics Summary Cards
        total_voters = Voter.objects.count()
        voted_count = Voter.objects.filter(vote__post__election=election).distinct().count()
        turnout_pct = (voted_count / total_voters * 100) if total_voters > 0 else 0.0

        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 18

        stats_headers = [
            ("A", "Total Registered Voters"),
            ("B", "Total Votes Cast"),
            ("C", "Voter Turnout %"),
            ("D", "Export Date & Time")
        ]

        # Merge for dates to span columns D to G
        ws.merge_cells('D4:G4')
        ws.merge_cells('D5:G5')

        # Style the headers range
        style_range(ws, 4, 1, 4, 3, font=font_stats_lbl, fill=fill_stats, alignment=align_center, border=thin_border)
        style_range(ws, 4, 4, 4, 7, font=font_stats_lbl, fill=fill_stats, alignment=align_center, border=thin_border)

        ws["A4"] = "Total Registered Voters"
        ws["B4"] = "Total Votes Cast"
        ws["C4"] = "Voter Turnout %"
        ws["D4"] = "Export Date & Time"

        ws["A5"] = total_voters
        ws["A5"].number_format = '#,##0'
        ws["B5"] = voted_count
        ws["B5"].number_format = '#,##0'
        ws["C5"] = f"{turnout_pct:.2f}%"
        ws["D5"] = generation_time

        style_range(ws, 5, 1, 5, 3, font=font_stats_val, fill=fill_stats, alignment=align_center, border=thin_border)
        style_range(ws, 5, 4, 5, 7, font=font_stats_val, fill=fill_stats, alignment=align_center, border=thin_border)

        # 3. Position Results Section
        current_row = 7
        posts = election.posts.all().prefetch_related('candidates')

        for post in posts:
            candidates = post.candidates.all()
            candidate_votes = Vote.objects.filter(post=post).values('candidate_id').annotate(vote_count=Count('id'))
            votes_map = {item['candidate_id']: item['vote_count'] for item in candidate_votes}
            total_post_votes = sum(votes_map.values())

            candidate_results = []
            for candidate in candidates:
                votes = votes_map.get(candidate.id, 0)
                percentage = (votes / total_post_votes) if total_post_votes > 0 else 0.0
                candidate_results.append({
                    'name': candidate.name,
                    'class': candidate._class,
                    'stream': candidate.stream,
                    'votes': votes,
                    'percentage': percentage,
                })

            # Sort by vote count descending, then name alphabetically
            candidate_results.sort(key=lambda x: (-x['votes'], x['name']))

            # Assign Rank and Status (Winner vs Runner-up)
            for rank, res in enumerate(candidate_results, start=1):
                res['rank'] = rank
                res['status'] = 'Winner' if rank <= post.required_selections and res['votes'] > 0 else ('Runner-up' if res['votes'] > 0 else '-')

            # Render Position Title Bar
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
            style_range(
                ws,
                current_row, 1, current_row, 7,
                font=font_section,
                fill=fill_section,
                alignment=align_left,
                border=Border(bottom=Side(border_style="medium", color="1F4E78"))
            )
            ws.cell(row=current_row, column=1).value = f"Position: {post.title} (Required Selections: {post.required_selections})"

            current_row += 1

            # Render Table Headers
            ws.row_dimensions[current_row].height = 20
            headers = ["Rank", "Candidate Name", "Class", "Stream", "Votes Received", "Percentage", "Result"]
            for col_idx, text in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = text
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            current_row += 1

            # Render Candidates list
            if not candidate_results:
                ws.row_dimensions[current_row].height = 18
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                style_range(
                    ws,
                    current_row, 1, current_row, 7,
                    font=Font(italic=True, color="7F7F7F"),
                    alignment=align_center,
                    border=thin_border
                )
                ws.cell(row=current_row, column=1).value = "No candidates registered for this position."
                current_row += 1
            else:
                for idx, res in enumerate(candidate_results):
                    ws.row_dimensions[current_row].height = 20
                    is_winner = res['status'] == 'Winner'
                    is_even = idx % 2 == 1

                    row_fill = fill_winner if is_winner else (fill_zebra if is_even else PatternFill(fill_type=None))

                    cells = [
                        (1, res['rank'], align_center),
                        (2, res['name'], align_left),
                        (3, res['class'], align_center),
                        (4, res['stream'], align_left),
                        (5, res['votes'], align_right),
                        (6, res['percentage'], align_right),
                        (7, res['status'], align_center)
                    ]

                    for col_idx, val, align in cells:
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = val
                        cell.font = font_winner if (is_winner and col_idx == 7) else font_candidate
                        if row_fill.fill_type:
                            cell.fill = row_fill
                        cell.alignment = align
                        cell.border = thin_border

                        # Formats
                        if col_idx == 5:
                            cell.number_format = '#,##0'
                        elif col_idx == 6:
                            cell.number_format = '0.0%'

                    current_row += 1

            # Add spacer rows between tables
            current_row += 2

        # Adjust Columns layout dimension widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@admin.register(Election)
class ElectionAdmin(admin.ModelAdmin):
    list_display = ('title', 'is_active', 'is_demo', 'created_at', 'export_results_link')
    list_filter = ('is_active', 'is_demo')
    search_fields = ('title',)
    actions = ['reset_election_votes', 'export_election_results_action']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:election_id>/export-results/', self.admin_site.admin_view(self.export_results), name='posts_election_export_results'),
        ]
        return custom_urls + urls

    def export_results(self, request, election_id):
        """View endpoint to export results for a single election."""
        try:
            election = Election.objects.get(pk=election_id)
        except Election.DoesNotExist:
            from django.contrib import messages
            self.message_user(request, "Election not found.", messages.ERROR)
            return redirect('..')

        output = generate_results_excel([election])
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="election_results_{election_id}.xlsx"'
        return response

    @admin.action(description="Export selected elections results to Excel")
    def export_election_results_action(self, request, queryset):
        """Admin action to export results for all selected elections."""
        if not queryset.exists():
            return

        output = generate_results_excel(list(queryset))
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = "election_results_multiple.xlsx" if queryset.count() > 1 else f"election_results_{queryset.first().id}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_results_link(self, obj):
        from django.urls import reverse
        from django.utils.html import format_html
        url = reverse('admin:posts_election_export_results', args=[obj.pk])
        return format_html('<a class="button" style="background-color: #2F5597; color: white; padding: 4px 10px; border-radius: 4px; text-decoration: none; font-weight: bold; font-size: 11px;" href="{}">Export Results</a>', url)
    export_results_link.short_description = 'Excel Results'

    @admin.action(description="Reset selected elections (delete all cast votes)")
    def reset_election_votes(self, request, queryset):
        deleted_count = 0
        for election in queryset:
            votes = Vote.objects.filter(post__election=election)
            count = votes.count()
            votes.delete()
            deleted_count += count
        self.message_user(request, f"Successfully reset {queryset.count()} election(s). Deleted {deleted_count} vote(s).")


@admin.register(Post)
class PostAdmin(admin.ModelAdmin):
    list_display = ('title', 'election', 'description', 'eligible_houses_display')
    list_filter = ('election',)
    search_fields = ('title', 'description')
    inlines = [EligibleHouseInline]

    def eligible_houses_display(self, obj):
        houses = obj.eligible_houses.all()
        if not houses:
            return 'All Houses'
        return ', '.join(h.get_house_display() for h in houses)
    eligible_houses_display.short_description = 'Eligible Houses'
